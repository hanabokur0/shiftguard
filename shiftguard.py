#!/usr/bin/env python3
"""
ShiftGuard - 労務リスク警告付き簡易シフト生成ツール
ローカル完結、ルールベースの最小実装
"""

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Tuple
import calendar

import pandas as pd
import yaml

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import jpholiday
except ImportError:
    jpholiday = None
    print("警告: jpholiday がインストールされていません。祝日判定が行われません。", file=sys.stderr)


class ShiftGuard:
    """シフト生成と労務リスク警告のコアクラス"""
    
    def __init__(self, rules_path: str):
        """
        Args:
            rules_path: ルール設定YAMLファイルのパス
        """
        with open(rules_path, 'r', encoding='utf-8') as f:
            self.rules = yaml.safe_load(f)
        
        self.staff_df = None
        self.config = {}
        self.schedule = []
        self.warnings = []
        
    def load_input(self, input_path: str):
        """入力Excelファイルを読み込む"""
        try:
            self.staff_df = pd.read_excel(input_path, sheet_name='staff')
            config_df = pd.read_excel(input_path, sheet_name='config')
            
            # configを辞書化
            for _, row in config_df.iterrows():
                self.config = row.to_dict()
                break  # 1行目のみ
            
            # データ検証
            self._validate_input()
            
        except Exception as e:
            raise ValueError(f"入力ファイルの読み込みに失敗しました: {e}")
    
    def _validate_input(self):
        """入力データの基本検証"""
        required_staff_cols = ['staff_id', 'name', 'desired_days', 'can_day', 
                                'can_night', 'can_weekend_holiday']
        for col in required_staff_cols:
            if col not in self.staff_df.columns:
                raise ValueError(f"staffシートに必須列 '{col}' がありません")
        
        required_config_keys = ['month', 'min_staff_day', 'min_staff_night']
        for key in required_config_keys:
            if key not in self.config or pd.isna(self.config[key]):
                raise ValueError(f"configに必須項目 '{key}' がありません")
    
    def generate_schedule(self):
        """シフトを生成する（ルールベース）"""
        print("シフト生成開始...")
        
        # 対象月の日付リストを作成（date型で統一）
        month_str = str(self.config['month'])
        year, month = map(int, month_str.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        
        # datetime ではなく date で統一（比較の安全性向上）
        from datetime import date
        dates = [date(year, month, day) for day in range(1, days_in_month + 1)]
        
        # ========== 供給チェック（枠モデル） ==========
        min_day = int(self.config.get('min_staff_day', 0))
        min_night = int(self.config.get('min_staff_night', 0))
        variable_extra = int(self.config.get('variable_extra_slots_month', 0))
        
        # 必要枠の計算
        base_slots = (min_day + min_night) * days_in_month
        total_slots_needed = base_slots + variable_extra
        
        # 供給枠の計算（各人の希望出勤日数の合計）
        supply_slots = sum(int(row['desired_days']) for _, row in self.staff_df.iterrows())
        
        print(f"枠モデル分析:")
        print(f"  ベース枠: {base_slots} （日勤{min_day}+夜勤{min_night} × {days_in_month}日）")
        print(f"  変動枠: {variable_extra}")
        print(f"  合計必要枠: {total_slots_needed}")
        print(f"  供給枠: {supply_slots} （全員の希望日数合計）")
        
        # 供給チェック警告
        if supply_slots < base_slots:
            self.warnings.append({
                'severity': 'RED',
                'code': 'INSUFFICIENT_CAPACITY_BASE',
                'message': f'供給枠が不足（必要{base_slots}、供給{supply_slots}）- シフトが回りません',
                'evidence': f'不足: {base_slots - supply_slots} 枠'
            })
        elif supply_slots < total_slots_needed:
            self.warnings.append({
                'severity': 'YELLOW',
                'code': 'INSUFFICIENT_CAPACITY_PEAK',
                'message': f'変動枠まで満たせない（必要{total_slots_needed}、供給{supply_slots}）- 繁忙時に耐えられません',
                'evidence': f'不足: {total_slots_needed - supply_slots} 枠'
            })
        else:
            margin = supply_slots - total_slots_needed
            print(f"  余力: {margin} 枠 ✓")
        
        # ========== 変動枠の配分（土日祝優先） ==========
        # 各日の必要人数を調整（土日祝に優先的に変動枠を配分）
        daily_demand = {}
        weekend_dates = [d for d in dates if self._is_weekend_or_holiday(d)]
        weekday_dates = [d for d in dates if not self._is_weekend_or_holiday(d)]
        
        # まず土日祝に配分（1枠ずつ）
        extra_distributed = 0
        cycle_index = 0
        while extra_distributed < variable_extra:
            if weekend_dates:
                target_date = weekend_dates[cycle_index % len(weekend_dates)]
                daily_demand[target_date] = daily_demand.get(target_date, 0) + 1
                extra_distributed += 1
                cycle_index += 1
                
                if extra_distributed >= variable_extra:
                    break
            
            # 土日祝だけで足りない場合は平日にも配分
            if cycle_index >= len(weekend_dates) * 2:  # 土日に2周したら平日へ
                if weekday_dates:
                    target_date = weekday_dates[extra_distributed % len(weekday_dates)]
                    daily_demand[target_date] = daily_demand.get(target_date, 0) + 1
                    extra_distributed += 1
                else:
                    break
        
        if variable_extra > 0:
            print(f"  変動枠配分: 土日祝{len([d for d in daily_demand if d in weekend_dates])}日、平日{len([d for d in daily_demand if d in weekday_dates])}日")
        
        # 各スタッフの状態を初期化
        staff_state = {}
        for _, staff in self.staff_df.iterrows():
            staff_id = staff['staff_id']
            
            # 希望休日をパース（date型で統一）
            requested_off = []
            if pd.notna(staff.get('requested_off_dates')):
                off_str = str(staff['requested_off_dates'])
                requested_off = [datetime.strptime(d.strip(), '%Y-%m-%d').date() 
                                for d in off_str.split(',') if d.strip()]
            
            staff_state[staff_id] = {
                'name': staff['name'],
                'desired_days': int(staff['desired_days']),
                'can_day': bool(staff['can_day']),
                'can_night': bool(staff['can_night']),
                'can_weekend_holiday': bool(staff['can_weekend_holiday']),
                'requested_off': requested_off,
                'assigned_days': 0,
                'last_shift_date': None,
                'last_shift_type': None,
                'consecutive_days': 0,
                'shifts': {}  # date -> shift_type
            }
        
        # フェーズ1: 希望休を先に埋める
        for date in dates:
            for staff_id, state in staff_state.items():
                if date in state['requested_off']:
                    state['shifts'][date] = 'OFF'
        
        # フェーズ2: 各日に必要人数を満たすよう割り当て
        for date in dates:
            is_weekend_holiday = self._is_weekend_or_holiday(date)
            
            # その日の割り当て状況
            day_assigned = []
            night_assigned = []
            
            # 既にOFFの人を除外
            available_staff = [sid for sid, state in staff_state.items() 
                              if state['shifts'].get(date) != 'OFF']
            
            # 優先順位でソート（土日祝不可の人は後回し、でも除外はしない）
            def priority_key(sid):
                state = staff_state[sid]
                # 希望日数に達していない人を優先
                shortage = state['assigned_days'] - state['desired_days']
                # 土日祝不可の人は優先度を下げる（+1000でペナルティ）
                if is_weekend_holiday and not state['can_weekend_holiday']:
                    shortage += 1000
                return shortage
            
            available_staff.sort(key=priority_key)
            
            # その日の必要人数（変動枠を上乗せ）
            extra_demand = daily_demand.get(date, 0)
            min_day_today = min_day + (extra_demand // 2)  # 変動枠を日勤・夜勤で半分ずつ
            min_night_today = min_night + (extra_demand - extra_demand // 2)  # 残りを夜勤に
            
            # 日勤割り当て
            for staff_id in available_staff:
                if len(day_assigned) >= min_day_today:
                    break
                state = staff_state[staff_id]
                if state['can_day'] and state['assigned_days'] < state['desired_days']:
                    # 休息時間チェック（簡易）
                    if self._can_assign_shift(state, date, 'DAY'):
                        state['shifts'][date] = 'DAY'
                        day_assigned.append(staff_id)
                        state['assigned_days'] += 1
                        self._update_state(state, date, 'DAY')
            
            # 夜勤割り当て
            for staff_id in available_staff:
                if staff_id in day_assigned:
                    continue
                if len(night_assigned) >= min_night_today:
                    break
                state = staff_state[staff_id]
                if state['can_night'] and state['assigned_days'] < state['desired_days']:
                    if self._can_assign_shift(state, date, 'NIGHT'):
                        state['shifts'][date] = 'NIGHT'
                        night_assigned.append(staff_id)
                        state['assigned_days'] += 1
                        self._update_state(state, date, 'NIGHT')
            
            # 人数不足警告（ワンオペ判定を含む）
            allow_solo_day = int(self.config.get('allow_solo_day', 0))
            allow_solo_night = int(self.config.get('allow_solo_night', 0))
            
            # 日勤
            if len(day_assigned) < min_day_today:
                # ワンオペ判定
                if len(day_assigned) == 1 and allow_solo_day:
                    # ワンオペ許可されているが警告は出す
                    self.warnings.append({
                        'severity': 'YELLOW',
                        'code': 'SOLO_SHIFT_DAY',
                        'message': f'日勤がワンオペ体制（構造リスク）',
                        'evidence': f'{date.strftime("%Y-%m-%d")}'
                    })
                else:
                    self.warnings.append({
                        'severity': 'RED',
                        'code': 'UNDERSTAFFED_DAY',
                        'message': f'日勤の必要人数不足 (必要: {min_day_today}, 実際: {len(day_assigned)})',
                        'evidence': f'{date.strftime("%Y-%m-%d")}'
                    })
            
            # 夜勤
            if len(night_assigned) < min_night_today:
                # ワンオペ判定
                if len(night_assigned) == 1 and allow_solo_night:
                    # ワンオペ許可されているが警告は出す
                    self.warnings.append({
                        'severity': 'YELLOW',
                        'code': 'SOLO_SHIFT_NIGHT',
                        'message': f'夜勤がワンオペ体制（構造リスク）',
                        'evidence': f'{date.strftime("%Y-%m-%d")}'
                    })
                else:
                    self.warnings.append({
                        'severity': 'RED',
                        'code': 'UNDERSTAFFED_NIGHT',
                        'message': f'夜勤の必要人数不足 (必要: {min_night_today}, 実際: {len(night_assigned)})',
                        'evidence': f'{date.strftime("%Y-%m-%d")}'
                    })
        
        # フェーズ4: 未割当の日を OFF で埋める（全員×全日の完全な表を作る）
        for staff_id, state in staff_state.items():
            for date in dates:
                if date not in state['shifts']:
                    state['shifts'][date] = 'OFF'
        
        # スケジュールをフラット化
        for staff_id, state in staff_state.items():
            for date in sorted(state['shifts'].keys()):
                shift_type = state['shifts'][date]
                self.schedule.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'shift_type': shift_type,
                    'staff_id': staff_id,
                    'name': state['name']
                })
        
        # フェーズ5: 労務リスクチェック
        self._check_labor_risks(staff_state)
        
        print(f"シフト生成完了: {len(self.schedule)} エントリ")
        print(f"警告: {len(self.warnings)} 件")
    
    def _is_weekend_or_holiday(self, d) -> bool:
        """土日または祝日かを判定（date型を受け取る）"""
        # date型に統一（datetimeが来ても対応）
        from datetime import date as date_type
        if hasattr(d, 'date'):
            d = d.date()
        
        # 土日判定
        if d.weekday() >= 5:  # 5=土, 6=日
            return True
        
        # 祝日判定
        if jpholiday:
            return jpholiday.is_holiday(d)
        
        return False
    
    def _can_assign_shift(self, state: dict, d, shift_type: str) -> bool:
        """シフト割り当て可能かチェック（休息時間など）"""
        from datetime import timedelta
        
        if state['last_shift_date'] is None:
            return True
        
        # 前日のシフトを確認
        prev_date = d - timedelta(days=1)
        if prev_date not in state['shifts']:
            return True
        
        prev_shift = state['shifts'][prev_date]
        
        # 簡易チェック: 夜勤→日勤は避ける（休息時間不足）
        # NOTE: 実際の休息時間計算は時刻情報が必要だが、ここでは簡易判定
        if prev_shift == 'NIGHT' and shift_type == 'DAY':
            return False
        
        return True
    
    def _update_state(self, state: dict, d, shift_type: str):
        """スタッフ状態を更新"""
        from datetime import timedelta
        
        state['last_shift_date'] = d
        state['last_shift_type'] = shift_type
        
        # 連続勤務日数カウント
        if shift_type != 'OFF':
            prev_date = d - timedelta(days=1)
            if prev_date in state['shifts'] and state['shifts'][prev_date] != 'OFF':
                state['consecutive_days'] += 1
            else:
                state['consecutive_days'] = 1
        else:
            state['consecutive_days'] = 0
    
    def _check_labor_risks(self, staff_state: dict):
        """労務リスクをチェック"""
        print("労務リスクチェック中...")
        
        # rules.yml から閾値を読む（config にない場合のデフォルト）
        max_consecutive_yellow = self.rules.get('thresholds', {}).get('max_consecutive_workdays', {}).get('yellow', 6)
        max_consecutive_red = self.rules.get('thresholds', {}).get('max_consecutive_workdays', {}).get('red', 8)
        max_overtime_yellow = self.rules.get('thresholds', {}).get('max_month_overtime_hours', {}).get('yellow', 45)
        max_overtime_red = self.rules.get('thresholds', {}).get('max_month_overtime_hours', {}).get('red', 54)
        
        # config から読む（実運用値）
        max_consecutive = int(self.config.get('max_consecutive_workdays', max_consecutive_yellow))
        standard_day_hours = float(self.config.get('standard_day_shift_hours', 8))
        standard_night_hours = float(self.config.get('standard_night_shift_hours', 10))
        max_overtime = float(self.config.get('max_month_overtime_hours', max_overtime_yellow))
        
        for staff_id, state in staff_state.items():
            # 1. 希望休違反チェック
            for req_date in state['requested_off']:
                if req_date in state['shifts'] and state['shifts'][req_date] != 'OFF':
                    self.warnings.append({
                        'severity': 'RED',
                        'code': 'REQUESTED_OFF_VIOLATION',
                        'message': f'希望休に勤務が割り当てられています',
                        'evidence': f'{state["name"]} ({req_date.strftime("%Y-%m-%d")})'
                    })
            
            # 2. 連続勤務日数チェック（rules の閾値を使用）
            consecutive = 0
            max_found = 0
            for date in sorted(state['shifts'].keys()):
                if state['shifts'][date] != 'OFF':
                    consecutive += 1
                    max_found = max(max_found, consecutive)
                else:
                    consecutive = 0
            
            if max_found > max_consecutive_red:
                self.warnings.append({
                    'severity': 'RED',
                    'code': 'EXCESSIVE_CONSECUTIVE',
                    'message': f'連続勤務日数が過剰 ({max_found}日、上限{max_consecutive_red}日)',
                    'evidence': f'{state["name"]}'
                })
            elif max_found > max_consecutive_yellow:
                self.warnings.append({
                    'severity': 'YELLOW',
                    'code': 'HIGH_CONSECUTIVE',
                    'message': f'連続勤務日数が多い ({max_found}日、推奨{max_consecutive_yellow}日)',
                    'evidence': f'{state["name"]}'
                })
            
            # 3. 休息時間違反チェック（簡易）
            sorted_dates = sorted([d for d, s in state['shifts'].items() if s != 'OFF'])
            for i in range(len(sorted_dates) - 1):
                curr_shift = state['shifts'][sorted_dates[i]]
                next_shift = state['shifts'][sorted_dates[i + 1]]
                
                if curr_shift == 'NIGHT' and next_shift == 'DAY':
                    if (sorted_dates[i + 1] - sorted_dates[i]).days == 1:
                        self.warnings.append({
                            'severity': 'RED',
                            'code': 'INSUFFICIENT_REST',
                            'message': f'休息時間不足の可能性 (夜勤→日勤)',
                            'evidence': f'{state["name"]} ({sorted_dates[i].strftime("%Y-%m-%d")} → {sorted_dates[i+1].strftime("%Y-%m-%d")})'
                        })
            
            # 4. 推定残業チェック（rules の閾値を使用）
            work_days = sum(1 for s in state['shifts'].values() if s != 'OFF')
            day_shifts = sum(1 for s in state['shifts'].values() if s == 'DAY')
            night_shifts = sum(1 for s in state['shifts'].values() if s == 'NIGHT')
            
            total_hours = day_shifts * standard_day_hours + night_shifts * standard_night_hours
            standard_hours = self.rules.get('standard_month_hours', 160)  # 月の標準労働時間
            estimated_overtime = max(0, total_hours - standard_hours)
            
            if estimated_overtime > max_overtime_red:
                self.warnings.append({
                    'severity': 'RED',
                    'code': 'EXCESSIVE_OVERTIME',
                    'message': f'推定残業時間が過剰 ({estimated_overtime:.1f}h、上限{max_overtime_red}h)',
                    'evidence': f'{state["name"]}'
                })
            elif estimated_overtime > max_overtime_yellow:
                self.warnings.append({
                    'severity': 'YELLOW',
                    'code': 'HIGH_OVERTIME',
                    'message': f'推定残業時間が上限に近い ({estimated_overtime:.1f}h、推奨{max_overtime_yellow}h)',
                    'evidence': f'{state["name"]}'
                })
            
            # 5. 土日祝不可なのに割り当てチェック
            if not state['can_weekend_holiday']:
                for date, shift in state['shifts'].items():
                    if shift != 'OFF' and self._is_weekend_or_holiday(date):
                        self.warnings.append({
                            'severity': 'YELLOW',
                            'code': 'WEEKEND_RESTRICTION',
                            'message': f'土日祝不可のスタッフに土日祝勤務',
                            'evidence': f'{state["name"]} ({date.strftime("%Y-%m-%d")})'
                        })
        
        # 警告がなければGREEN
        if not self.warnings:
            self.warnings.append({
                'severity': 'GREEN',
                'code': 'ALL_CLEAR',
                'message': '重大な労務リスクは検出されませんでした',
                'evidence': ''
            })
    
    def save_output(self, output_path: str):
        """結果をExcelに出力"""
        print(f"結果を保存中: {output_path}")
        
        # スケジュールをDataFrameに
        schedule_df = pd.DataFrame(self.schedule)
        if not schedule_df.empty:
            schedule_df = schedule_df.sort_values(['date', 'shift_type', 'staff_id'])
        
        # 警告をDataFrameに
        warnings_df = pd.DataFrame(self.warnings)
        
        # Excel出力
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1) 縦持ち（分析向け）
            schedule_df.to_excel(writer, sheet_name='schedule', index=False)
            warnings_df.to_excel(writer, sheet_name='warnings', index=False)# 2) カレンダー見た目（現場向け）：名前 × 日付 の横持ち（曜日行 + 色分け）
if not schedule_df.empty:
    # ---- 2-1. データを横持ちに変換（name × date） ----
    calendar_df = schedule_df.pivot(index='name', columns='date', values='shift_type')
    calendar_df = calendar_df.reindex(sorted(calendar_df.columns), axis=1)

    # 表示用の記号に置換（必要ならここでカスタム）
    display_map = {'DAY': 'D', 'NIGHT': 'N', 'OFF': '休'}
    calendar_df = calendar_df.fillna('休').replace(display_map)

    # ---- 2-2. まずDataFrameとして書き出し（開始位置をずらす）----
    # レイアウト：A列=名前、B列以降=日付
    # 上部に「年月」「日付行」「曜日行」を入れたいので、開始行を3行下げる
    start_row = 3  # 0-indexではなくExcelの行番号（1始まり）で扱うため後で+1調整する
    sheet_name = 'calendar'
    calendar_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=0)

    # ---- 2-3. openpyxlで体裁を整える ----
    wb = writer.book
    ws = writer.sheets[sheet_name]

    # 日付リスト（YYYY-MM-DD）
    date_cols = list(calendar_df.columns)

    # 年月タイトル（configのmonthを優先）
    ym = str(self.config.get('month', ''))
    if not ym and date_cols:
        ym = date_cols[0][:7]

    title = f"{ym} 勤務表"
    # タイトル行（1行目）を作る：A1〜最終列を結合
    header_row_date = start_row + 1  # 日付行（Excel上）
    header_row_wday = start_row + 2  # 曜日行（Excel上）
    title_row = 1

    # 最終列（A=1, B=2 ...）
    last_col = 1 + len(date_cols)  # A列=名前 + 日付列数
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_col)
    cell = ws.cell(row=title_row, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日付行・曜日行を上書きする（DataFrameのヘッダは start_row+1 に入っている）
    # A列ヘッダ（name）を「メンバー」にする
    ws.cell(row=header_row_date, column=1, value="メンバー")
    ws.cell(row=header_row_wday, column=1, value="")

    # 罫線・色
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_sat = PatternFill("solid", fgColor="D9EAF7")   # 土：薄青
    fill_sun = PatternFill("solid", fgColor="F9D6D5")   # 日祝：薄赤
    fill_week = PatternFill("solid", fgColor="FFFFFF")  # 平日
    fill_off = PatternFill("solid", fgColor="EEEEEE")   # 休：薄グレー
    fill_day = PatternFill("solid", fgColor="E6F4EA")   # D：薄緑
    fill_night = PatternFill("solid", fgColor="E8F0FE") # N：薄青紫

    font_header = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 日付行（header_row_date）と曜日行（header_row_wday）をセット
    for i, dstr in enumerate(date_cols, start=2):  # B列から
        # 表示：M/D
        try:
            d = datetime.strptime(dstr, "%Y-%m-%d").date()
        except Exception:
            d = None

        date_label = f"{d.month}/{d.day}" if d else dstr
        ws.cell(row=header_row_date, column=i, value=date_label)

        # 曜日/祝日
        if d:
            # 祝日判定
            is_holiday = bool(jpholiday and jpholiday.is_holiday(d))
            w = "月火水木金土日"[d.weekday()]
            wlabel = "祝" if is_holiday else w
            ws.cell(row=header_row_wday, column=i, value=wlabel)

            # 列色（曜日ヘッダの背景）
            if is_holiday or d.weekday() == 6:
                col_fill = fill_sun
            elif d.weekday() == 5:
                col_fill = fill_sat
            else:
                col_fill = fill_week
        else:
            ws.cell(row=header_row_wday, column=i, value="")
            col_fill = fill_week

        for r in (header_row_date, header_row_wday):
            c = ws.cell(row=r, column=i)
            c.font = font_header
            c.alignment = align_center
            c.fill = col_fill
            c.border = border

    # A列ヘッダの体裁
    for r in (header_row_date, header_row_wday):
        c = ws.cell(row=r, column=1)
        c.font = font_header
        c.alignment = align_center
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.border = border

    # 本体セルの体裁（勤務記号で色）
    body_start_row = start_row + 2 + 1  # 曜日行の次の行（DataFrameの1行目）
    body_end_row = body_start_row + len(calendar_df.index) - 1
    body_start_col = 1
    body_end_col = last_col

    for r in range(body_start_row, body_end_row + 1):
        for ccol in range(body_start_col, body_end_col + 1):
            cell = ws.cell(row=r, column=ccol)
            cell.alignment = align_center
            cell.border = border

            # 名前列は薄いグレー
            if ccol == 1:
                cell.fill = PatternFill("solid", fgColor="F7F7F7")
                continue

            v = str(cell.value) if cell.value is not None else ""
            if v == "休":
                cell.fill = fill_off
            elif v == "D":
                cell.fill = fill_day
            elif v == "N":
                cell.fill = fill_night

    # 列幅調整
    ws.column_dimensions["A"].width = 14
    for i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 4

    # 行高
    ws.row_dimensions[title_row].height = 22
    ws.row_dimensions[header_row_date].height = 18
    ws.row_dimensions[header_row_wday].height = 18

    # ウィンドウ枠固定（タイトルは固定しない：日付/曜日/名前を固定）
    ws.freeze_panes = ws["B{}".format(body_start_row)]

    # 印刷設定（横向き、1ページに収めやすく）
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
        
        print("保存完了")
    
    def print_summary(self):
        """サマリーを表示"""
        print("\n" + "=" * 60)
        print("生成結果サマリー")
        print("=" * 60)
        
        # 警告レベル別集計
        severity_counts = {}
        for w in self.warnings:
            sev = w['severity']
            severity_counts[sev] = severity_counts.get(sev, 0) + 1
        
        print(f"\n警告件数:")
        for sev in ['RED', 'YELLOW', 'GREEN']:
            count = severity_counts.get(sev, 0)
            if count > 0:
                print(f"  {sev}: {count}件")
        
        # RED警告を表示
        red_warnings = [w for w in self.warnings if w['severity'] == 'RED']
        if red_warnings:
            print(f"\n【重大な問題 (RED)】")
            for w in red_warnings[:5]:  # 最大5件
                print(f"  - {w['message']}: {w['evidence']}")
            if len(red_warnings) > 5:
                print(f"  ... 他 {len(red_warnings) - 5} 件")


def main():
    parser = argparse.ArgumentParser(
        description='ShiftGuard - 労務リスク警告付き簡易シフト生成ツール'
    )
    parser.add_argument('--input', '-i', required=True, help='入力Excelファイル')
    parser.add_argument('--output', '-o', required=True, help='出力Excelファイル')
    parser.add_argument('--rules', '-r', default='rules.yml', help='ルール設定ファイル (default: rules.yml)')
    
    args = parser.parse_args()
    
    # 入力ファイル存在チェック
    if not Path(args.input).exists():
        print(f"エラー: 入力ファイルが見つかりません: {args.input}", file=sys.stderr)
        return 1
    
    if not Path(args.rules).exists():
        print(f"エラー: ルールファイルが見つかりません: {args.rules}", file=sys.stderr)
        return 1
    
    try:
        # シフト生成実行
        guard = ShiftGuard(args.rules)
        guard.load_input(args.input)
        guard.generate_schedule()
        guard.save_output(args.output)
        guard.print_summary()
        
        print(f"\n✓ 処理完了: {args.output}")
        return 0
        
    except Exception as e:
        print(f"エラー: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
